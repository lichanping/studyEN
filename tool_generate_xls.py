import asyncio
import os
import random
import re
import time
from datetime import datetime
from os.path import dirname, abspath

import edge_tts
import numpy as np
import pandas as pd
from edge_tts import VoicesManager
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from ptest.decorator import TestClass, Test
from bs4 import BeautifulSoup


def get_sub_folder_path(sub_dir_name='user_data'):
    """
    Create the destination folder if not exists.
    :param sub_dir_name: default is 'data'
    :return: sub folder's absolute path.
    """
    current_dir_name = dirname(__file__)
    abs_path = abspath(current_dir_name)
    sub_folder = os.sep.join([abs_path, sub_dir_name])
    return sub_folder


class TxtToXLSX:
    def __init__(self):
        self.data_folder = get_sub_folder_path()
        self.sound_folder = get_sub_folder_path('sounds')
        self.ori_file = None
        self.generate_file = None
        self.missing_words = []

    def convert(self, file_name):
        extracted_data = self.read_text(file_name)
        # self.create_excel(extracted_data)

    def remove_duplicates_or_merge_translations(self, file_name):
        """
        Remove duplicate English words or merge their translations directly from the original text file.
        """
        file_path = os.path.join(self.data_folder, file_name)
        english_words = {}  # Dictionary to store English words and their translations
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                match = re.match(r'([a-zA-Z\'\s\-\.\/]+)\s*(.*)', line.strip())
                if match:
                    english_word, translation = match.groups()
                    english_word = english_word.strip()

                    if english_word.endswith(('adj.', 'adv.', 'n.', 'v.', 'phr.', 'vt.', 'prep.', 'vi.', 'det.',
                                              'pron.', 'conj.', 'int.', 'aux.', 'a.', 'ad.', 'n./ad.', 'num.',
                                              'a./ pron.', 'adv', 'n．', 'art.')):
                        # If it does, move the part of speech to the translation
                        pos = english_word.split()[-1]  # Get the last part of the word as part of speech
                        english_word = english_word[
                                       :-len(pos)].strip()  # Remove the part of speech from the English word
                        translation = f"({pos}) {translation.strip()}"

                    # Replace "sb" with "somebody" and "sth" with "something" only in the English words part
                    if english_word not in ['posthumous', 'aesthetic', 'disbursement']:
                        english_word = english_word.replace("sb", "somebody").replace("sth", "something")
                    english_word = re.sub(r'sw(?!\w)', 'somewhere', english_word)
                    translation = translation.strip()
                    if english_word not in english_words:
                        # If the English word is encountered for the first time, initialize its translations as a list
                        english_words[english_word] = [translation]
                    else:
                        # If the English word already exists, append the new translation to its list of translations
                        existing_translation = english_words[english_word]
                        if translation not in existing_translation:
                            existing_translation.append(translation)

        # Write the unique English words and their translations back to the original file
        with open(file_path, 'w', encoding='utf-8') as file:
            for english_word, translations in english_words.items():
                # Merge translations into a single string, separated by semicolons
                merged_translations = ';'.join(translations)
                file.write(f"{english_word}\t{merged_translations}\n")

    def read_text(self, file_name):
        """
        Generate MissingSound.txt if audio not exists
        :param file_name:
        :return:
        """
        self.ori_file = file_name
        self.generate_file = os.path.join(self.data_folder, file_name.split('.')[0] + "_抗遗忘单词.xlsx")
        missing_sound_file = os.path.join(self.data_folder, "MissingSound.txt")  # Path to store missing sound words
        file_path = os.path.join(self.data_folder, file_name)
        data = []
        missing_words = []  # List to store missing sound words
        pattern = re.compile(r'([a-zA-Z\'\’\s\-\.\/]+)\s*(.*)')
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                match = pattern.match(line.strip())
                if match:
                    english_word, translation = match.groups()
                    english_word = english_word.strip()
                    # Replace "sb" with "somebody" and "sth" with "something" only in the English words part
                    english_word = english_word.replace("sb", "somebody").replace("sth", "something")
                    translation = translation.strip()
                    media = os.path.join(self.sound_folder, f"{english_word.strip()}.mp3")
                    exist = os.path.exists(media)
                    # print(f"English: {english_word}, Translation: {translation}, Sound: {exist}")
                    data.append({"单词": english_word, "释意": translation, "音频": str(exist)})
                    if not exist:
                        self.missing_words.append(english_word.strip())
                else:
                    print(f"Invalid format in line: {line.strip()}")
        # Write missing words to MissingSound.txt
        with open(missing_sound_file, 'w', encoding='utf-8') as missing_file:
            missing_file.write("\n".join(self.missing_words))

        return data

    def create_excel(self, data):
        df = pd.DataFrame(data)
        df.insert(0, '序号', range(1, len(df) + 1))
        df.to_excel(self.generate_file, index=False)
        print(f"Excel file '{self.generate_file}' created successfully.")


class TextToSpeechConverter:
    def __init__(self, txt_to_xlsx):
        self.txt_to_xlsx = txt_to_xlsx

    async def convert_text_to_audio(self, file_name, repeat=2, max_items=None):
        extracted_data = self.txt_to_xlsx.read_text(file_name)

        if max_items is not None and 0 < max_items < len(extracted_data):
            extracted_data = random.sample(extracted_data, max_items)

        current_date = datetime.now().strftime('%Y-%m-%d')
        output_file = os.path.join(self.txt_to_xlsx.data_folder, file_name.split('.')[0] + f"-{current_date}.mp3")

        voices = await VoicesManager.create()
        voice_names = [
            "Microsoft Server Speech Text to Speech Voice (en-US, AvaMultilingualNeural)",
            "Microsoft Server Speech Text to Speech Voice (en-US, EmmaMultilingualNeural)",
            "Microsoft Server Speech Text to Speech Voice (en-US, EmmaNeural)",
            "Microsoft Server Speech Text to Speech Voice (en-US, MichelleNeural)"
        ]
        english_voice = voice_names[-1]
        # english_voice = voices.find(Gender="Female", Language="en")
        # english_voice = random.choice(english_voice)["Name"]
        chinese_voice = voices.find(
            Language='zh'
            , Gender="Male"
            , Locale="zh-CN"
        )

        with open(output_file, "wb") as file:
            current_date = datetime.now().strftime('%Y-%m-%d')
            print(f"测验时间: ___{current_date}___, 考核项：___重点语言点___")
            for index, item in enumerate(extracted_data):
                english_word = item['单词']
                chinese_meaning = item['释意']
                # print(f"English: {english_word}, Translation: {chinese_meaning}")
                # print(f"{index + 1}: {english_word}, 翻译为_______")
                print(f"{index + 1}: _______, 翻译为 {chinese_meaning}")

                english_voice_name = english_voice
                chinese_voice_name = random.choice(chinese_voice)["Name"]

                # Repeat English audio twice
                for _ in range(repeat):
                    english_stream = edge_tts.Communicate(english_word, voice=english_voice_name).stream()
                    async for chunk in english_stream:
                        if chunk["type"] == "audio":
                            file.write(chunk["data"])

                chinese_stream = edge_tts.Communicate(chinese_meaning, voice=chinese_voice_name).stream()

                async for chunk in chinese_stream:
                    if chunk["type"] == "audio":
                        file.write(chunk["data"])

        print(f"Audio file '{output_file}' created successfully.")


@TestClass(run_mode='singleline')
class GenerateTool:
    @Test()
    def simplify_words(self):
        # remove duplicate words
        tool = TxtToXLSX()
        tool.remove_duplicates_or_merge_translations('雅思全部.txt')
        tool.remove_duplicates_or_merge_translations('高中考纲单词.txt')
        tool.remove_duplicates_or_merge_translations('高中考纲词组.txt')
        tool.remove_duplicates_or_merge_translations('敏珺语言点.txt')
        tool.remove_duplicates_or_merge_translations('中考作文高频词汇.txt')
        tool.remove_duplicates_or_merge_translations('中考词汇新增.txt')

    @Test()
    def calculate_missing_words(self):
        tool = TxtToXLSX()
        # generate missing sounds
        tool.convert('雅思全部.txt')
        tool.convert('高中考纲单词.txt')  # commented the create_excel due to uselessness.
        tool.convert('高中考纲词组.txt')
        tool.convert('敏珺语言点.txt')

    @Test()
    def generate_media_word_list(self):
        def en_and_cn(file, max_items):
            start_time = time.time()  # Record start time
            converter = TextToSpeechConverter(tool)
            asyncio.run(converter.convert_text_to_audio(file, max_items=max_items))
            end_time = time.time()  # Record end time
            elapsed_time = end_time - start_time  # Calculate elapsed time
            print(f"Time taken: {elapsed_time} seconds")

        tool = TxtToXLSX()
        en_and_cn('敏珺语言点.txt', max_items=None)
        # en_and_cn('高中考纲词组.txt', max_items=10)

    @Test()
    def calculate_class_fee(self):
        temp_folder = get_sub_folder_path('temp')
        file_path = os.path.join(temp_folder, '6月练习记录.xlsx')
        columns_to_remove = ['30分钟软件时长', '60分钟软件时长', '软件体验时长', '佣金总额', '身份证号码', '陪练手机号']
        self.add_category_sheet_with_fee(file_path, columns_to_remove)
        self.separate_classes_into_sheets(file_path)
        self.create_and_save_pivot_table(file_path)

    @Test()
    def calc_personal_fee(self):
        temp_folder = get_sub_folder_path('temp')
        file_path = os.path.join(temp_folder, '悠然-历史.html')
        self.calc_month_fee(file_path)

        file_path = os.path.join(temp_folder, '悠然-最近.html')
        self.calc_month_fee(file_path)

    def calc_month_fee(self, file_path):
        # Open the HTML file
        with open(file_path, "r", encoding="utf-8") as file:
            html_content = file.read()
        # Parse the HTML content
        soup = BeautifulSoup(html_content, 'html.parser')
        # Find the root node <ion-list> with class containing 'md'
        root_node = soup.find('ion-list', class_='md')
        # Get all child nodes <ion-item-sliding> with class "md"
        child_nodes = root_node.find_all('ion-item-sliding', class_='md')
        data = []
        reading_keywords = ['阅读理解', '文化阅读', '阅读真题', '语法', '完型填空']

        # Function to determine the category
        def get_category(course, class_time):
            if class_time == '体验课':
                return '体验课'
            elif any(keyword in course for keyword in reading_keywords):
                return '阅读完型语法课'
            else:
                return '词汇课'

        # Simplified function to determine the actual price
        def get_actual_price(class_time, category):
            if class_time == '体验课':
                return 40

            # Check for 60 or 30 minute classes
            duration = 60 if '60分钟' in class_time else 30 if '30分钟' in class_time else None
            if duration:
                base_price = 50 if category == '词汇课' else 55
                return base_price if duration == 60 else base_price / 2

            return 0  # Default if no condition is met

        for child in child_nodes:
            # Extract the <ion-item> node
            ion_item = child.find('ion-item')

            if ion_item:
                # Extract the <ion-label> node
                ion_label = ion_item.find('ion-label')

                if ion_label:
                    # Extract the student name (text in <h2>, excluding <span>)
                    h2_tag = ion_label.find('h2')
                    student = h2_tag.contents[0].strip()  # Get only the first part of the <h2> before any <span>

                    # Extract the class time (text in the first <span> under <h2>)
                    class_time = ion_label.find('h2').find('span').get_text(strip=True)

                    # Extract the first <p> text (课程)
                    first_p = ion_label.find('p')
                    course = first_p.get_text(strip=True) if first_p else ""

                    # Extract the last <p> text (价格)
                    last_p = ion_label.find_all('p')[-1]
                    price = last_p.get_text(strip=True) if last_p else ""

                    # Determine category based on course name and class time
                    category = get_category(course, class_time)

                    # Determine actual price based on class time and category
                    actual_price = get_actual_price(class_time, category)

                    # Append the extracted data to the list
                    data.append({
                        '学生': student,
                        '课时': class_time,
                        '课程': course,
                        '价格': price,
                        '类别': category,
                        '实际价格': actual_price
                    })
        # Create a DataFrame using the extracted data
        df = pd.DataFrame(data)
        # Create the pivot table
        df['时长'] = df['课时'].apply(lambda x: 1 if '60分钟' in x else 0.5 if '30分钟' in x else 1)
        pivot = pd.pivot_table(df, values='时长', index='类别', aggfunc='sum').reset_index()
        # Define the Excel file path
        temp_folder = get_sub_folder_path('temp')
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_file = os.path.join(temp_folder, f'{base_name}.xlsx')

        # Write the data and pivot table to a new Excel file, overriding if it exists
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='明细', index=False)
            pivot.to_excel(writer, sheet_name='汇总', index=False)
        print(f"Data and summary successfully written to {output_file}")

    def create_and_save_pivot_table(self, input_file_path):
        # 读取 Excel 文件中的数据
        df = pd.read_excel(input_file_path, sheet_name='总表')

        # 创建数据透视表
        pivot_table = pd.pivot_table(
            df,
            values=['应发放佣金（不包括成交奖励等）', '课时'],
            index=['陪练账号', '陪练姓名'],
            columns=['陪练类型', '课程类别'],
            aggfunc={'应发放佣金（不包括成交奖励等）': 'sum', '课时': 'sum'},
            fill_value=0  # 可选：填充缺失值
        )

        # 加载现有的 Excel 工作簿
        book = load_workbook(input_file_path)
        sheet_name = '数据透视表'
        if sheet_name in book.sheetnames:
            book.remove(book[sheet_name])

        # 使用 Pandas 的 ExcelWriter 并且加载现有工作簿
        with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a') as writer:
            writer.book = book
            pivot_table.to_excel(writer, sheet_name=sheet_name)

        print(f'数据透视表已添加到 {input_file_path}')

    def add_category_sheet_with_fee(self, file_path, columns_to_remove):
        # Load the workbook
        book = load_workbook(file_path)

        # Check if '类别' sheet exists, remove it if it does
        category_sheet_name = '总表'
        # Remove all sheets except 'sheet1'
        for sheet_name in book.sheetnames:
            if sheet_name != 'sheet1':
                book.remove(book[sheet_name])

        # Read the 'removed' sheet to calculate '课程类别' and 'fee'
        df = pd.read_excel(file_path, sheet_name='sheet1')
        # Remove specified columns from the dataframe
        df_removed = df.drop(columns=columns_to_remove)

        # Add a new column '课程类别' based on '资料名称'
        df_removed['课程类别'] = df_removed['资料名称'].apply(
            lambda x: '阅读完型语法课' if any(word in x for word in ['阅读理解', '文化阅读', '阅读真题', '语法', '完型填空']) else '词汇课')

        # Add a new column '课时' based on '陪练类型'
        df_removed['课时'] = df_removed['陪练类型'].apply(
            lambda x: 0.5 if '30分钟课' in x else 1)

        # Add a new column 'fee' based on '陪练类型' and '课程类别'
        def calculate_fee(row):
            if row['陪练类型'] == '体验课':
                return 40
            else:
                duration = 60 if '60分钟课' in row['陪练类型'] else 30
                if row['课程类别'] == '词汇课':
                    return 50 if duration == 60 else 25
                else:
                    return 55 if duration == 60 else 27.5

        df_removed['应发放佣金（不包括成交奖励等）'] = df_removed.apply(calculate_fee, axis=1)

        # Save the modified dataframe to a new sheet named '类别'
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            writer.book = book
            df_removed.to_excel(writer, index=False, sheet_name=category_sheet_name)

        print(f"New sheet '{category_sheet_name}' added after 'removed' sheet in the original file {file_path}")

    def separate_classes_into_sheets(self, file_path):
        # Load the workbook
        book = load_workbook(file_path)

        # Check if '体验课' and '正课' sheets exist, remove them if they do
        experience_sheet_name = '体验课'
        regular_class_sheet_name = '正课'
        if experience_sheet_name in book.sheetnames:
            book.remove(book[experience_sheet_name])
        if regular_class_sheet_name in book.sheetnames:
            book.remove(book[regular_class_sheet_name])

        # Read the '总表' sheet to filter records
        df_category = pd.read_excel(file_path, sheet_name='总表')

        # Filter the records
        df_experience = df_category[df_category['陪练类型'] == '体验课']
        df_regular = df_category[df_category['陪练类型'] != '体验课']

        # Save the filtered dataframes to new sheets named '体验课' and '正课'
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            writer.book = book
            df_experience.to_excel(writer, index=False, sheet_name=experience_sheet_name)
            df_regular.to_excel(writer, index=False, sheet_name=regular_class_sheet_name)

        print(
            f"Filtered data saved as new sheets '{experience_sheet_name}' and '{regular_class_sheet_name}' in the original file {file_path}")
